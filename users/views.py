from django.contrib.auth.decorators import user_passes_test
from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
from django.core.exceptions import ObjectDoesNotExist


from django.contrib.auth import authenticate, login
from django.shortcuts import render, redirect
from django.core.exceptions import ObjectDoesNotExist
from inscripcion.models import Grupo, Inscripcion, Asignatura, Periodo


from django.shortcuts import render, redirect
from django.contrib.auth import authenticate, login
from django.contrib import messages
from django.core.exceptions import ObjectDoesNotExist

from . forms import Carga_alumnos
import openpyxl
from django.contrib.auth.hashers import make_password
from .models import User
from django.shortcuts import render
from inscripcion.models import Inscripcion
from django.contrib.auth.models import Group
from django.db import IntegrityError


def login_users(request):
    if request.method == 'POST':
        numero_cuenta = request.POST['numero_cuenta']
        password = request.POST['password']
        user = authenticate(request, username=numero_cuenta, password=password)

        if user is not None:
            login(request, user)

            # try:
            #     # Intentar obtener la instancia de Inscripcion del usuario
            #     inscripcion = user.alumno
            # except ObjectDoesNotExist:
            #     # Si no existe la instancia de Inscripcion, crearla
            #     perio = 'Sin periodo'
            #     inscripcion = Inscripcion(numero_cuenta=user)
            #     inscripcion.save()

            # Verificar si el usuario es administrativo
            if user.groups.filter(name='Administrativos').exists():
                return redirect('usuarios_inscritos_grupo')

            return redirect('index')
        else:
            messages.error(request, 'Número de Cuenta o contraseña inválidos')

    return render(request, 'authenticate/login_view.html', {'messages': messages.get_messages(request)})


def logout_users(request):
    logout(request)
    # Lógica adicional después del cierre de sesión
    return redirect('login')


def es_administrador(user):
    return user.is_authenticated and user.is_superuser


@user_passes_test(es_administrador)
def vista_administrador(request):
    # Código de la vista para el administrador aquí
    return render(request, 'authenticate/inscripcion.html', {'messages': messages.get_messages(request)})
    

def carga_users(request):
    periodos = Periodo.objects.all()  # Obtener todos los períodos disponibles

    if request.method == 'POST':
        periodo_id = request.POST.get('periodo')
        form = Carga_alumnos(request.POST, request.FILES)
        if form.is_valid():
            try:
                archivo = openpyxl.load_workbook(request.FILES['file'])
                grupo_alumnos = Group.objects.get(name='Alumnos')  # Obtener el grupo "Alumnos"
                periodo = Periodo.objects.get(activo = True)  # Obtener el período seleccionado


                for hoja in archivo.worksheets:  # Iterar sobre todas las hojas
                    for fila in hoja.iter_rows(min_row=2, values_only=True):  # Iterar sobre todas las filas
                        
                        
                        semestre = {
                            '1': 1,
                            '2': 2,
                            '3': 3,
                            '4': 4,
                            '5': 5,
                            '6': 6,
                            '7': 7,
                            '8': 8,
                            '9': 9,
                        }

                        semestre_actual = next((semestre[key] for key in semestre if key in hoja.title), None)
                        if semestre_actual is None:
                            semestre_actual = 1
                            continue  # Si no se encuentra un semestre válido, pasar a la siguiente hoja

                        grupo = fila[5]

                        if fila[2] is not None:
                            numero_cuenta = fila[1]
                            nombre_completo = fila[2]
                            
                            # Verificar si el usuario ya existe por número de cuenta
                            if User.objects.filter(numero_cuenta=numero_cuenta).exists(): 
                                messages.warning(request, f'El usuario con número de cuenta {numero_cuenta} {nombre_completo} ya tiene una cuenta.')
                                continue
                            
                            # Verificar que nombre_completo sea una cadena antes de intentar dividirlo
                            if not isinstance(nombre_completo, str):
                                messages.error(request, f'Error en la fila: {fila}. El nombre completo no es un nombre válido.')
                                continue
                            
                            # Separar el nombre completo en nombre y apellidos
                            nombre_partes = nombre_completo.split()
                            if len(nombre_partes) < 2:
                                apellido_paterno = nombre_partes[0] 
                                apellido_materno = ''
                                nombre_final = nombre_partes[0] if nombre_partes else ''
                            elif len(nombre_partes) == 2:
                                nombre_final = nombre_partes[0]
                                apellido_paterno = nombre_partes[1]
                                apellido_materno = ''
                            else:
                                nombre_final =  ' '.join(nombre_partes[2:])
                                apellido_paterno = nombre_partes[0]
                                apellido_materno = nombre_partes[1]
                            
                            # Generar un username único
                            base_username = '_'.join(nombre_partes)
                            username = base_username
                            contador = 1
                            while User.objects.filter(username=username).exists():
                                username = f"{base_username}_{contador}"
                                contador += 1
                            
                            
                            is_superuser = False
                            is_staff = False
                            is_active = True
                            email = fila[3] if fila[3] else 'default@example.com'
                            first_name = nombre_final
                            last_name = f"{apellido_paterno} {apellido_materno}"
                            
                            # Crear y guardar la instancia del modelo User
                            if numero_cuenta is not None:
                                contrasena_inicial = f"{numero_cuenta}{apellido_paterno}"
                                usuario = User.objects.create(
                                    numero_cuenta=numero_cuenta,
                                    is_superuser=is_superuser,
                                    username=username,
                                    first_name=first_name,
                                    last_name=last_name,
                                    email=email,
                                    is_staff=is_staff,
                                    is_active=is_active,
                                    semestre_actual=semestre_actual
                                )
                                usuario.set_password(contrasena_inicial)  # Establecer la contraseña como el número de cuenta seguido del apellido paterno
                                print(contrasena_inicial)
                                usuario.save()

                                # Asignar el usuario al grupo "Alumnos"
                                usuario.groups.add(grupo_alumnos)

                                # Crear la inscripción y asignar el período
                                inscripcion = Inscripcion.objects.create(
                                    numero_cuenta=usuario,
                                    periodo=periodo
                                )

                                
                                


                                messages.success(request, f'El usuario {first_name} {last_name} ha sido registrado exitosamente con la contraseña inicial: {contrasena_inicial}')

                # Procesar materias e inscripciones por cada alumno
                alumnos_materias_grupos = []
                materias_actuales = []
                grupos = []
                

                # Construir el mapeo de alumnos a sus materias desde el archivo
                for hoja in archivo.worksheets:  # Iterar sobre todas las hojas
                    for fila in hoja.iter_rows(min_row=2, values_only=True):
                        if fila[1] is not None:
                            if materias_actuales:
                                alumnos_materias_grupos.append((numero_cuenta, materias_actuales, grupos))
                            numero_cuenta = fila[1]
                            materias_actuales = []
                            grupos = []  # Reiniciar la lista de grupos para el siguiente alumno
                            
                        clave_asignatura = fila[3]
                        if clave_asignatura is not None:
                            materias_actuales.append(clave_asignatura)

                        clave_grupo = fila[5]
                        if clave_grupo is not None:
                            try:
                                clave_grupo = int(clave_grupo)  # Intentar convertir a entero
                            except ValueError:
                                pass  # Si no se puede convertir, mantener el valor original
                            if clave_grupo not in grupos:
                                grupos.append(clave_grupo)
                        

                # Añadir la última lista de materias si no está vacía
                if materias_actuales:
                    alumnos_materias_grupos.append((numero_cuenta, materias_actuales, grupos))
                            
                print(f"Alumnos y materias: {alumnos_materias_grupos}")



                
                
            


                # Registrar las inscripciones
                for numero_cuenta, materias, grupos in alumnos_materias_grupos:
                    try:
                        user = User.objects.get(numero_cuenta=numero_cuenta)
                        inscripcion, created = Inscripcion.objects.get_or_create(numero_cuenta=user, periodo=periodo)
                        for clave_asignatura in materias:
                            asignatura = Asignatura.objects.get(clave_asignatura=clave_asignatura)
                            inscripcion.asignatura.add(asignatura)
                        for clave_grupo in grupos:
                            try:
                                grupo = Grupo.objects.get(clave_grupo=clave_grupo)
                                inscripcion.grupo.add(grupo)
                            except Grupo.DoesNotExist:
                                messages.error(request, f'Error: El grupo con clave {clave_grupo} no existe.')
                         


                        messages.success(request, f'Materias inscritas para el alumno {user.first_name} {user.last_name}.')
                    except User.DoesNotExist:
                        messages.error(request, f'Error: El usuario con número de cuenta {numero_cuenta} no existe.')
                    except Asignatura.DoesNotExist:
                        messages.error(request, f'Error: La asignatura con clave_asignatura {clave_asignatura} no existe.')

                print("Materias inscritas para cada alumno.")
                
            except Exception as e:
                messages.error(request, f'Error al cargar el alumno: {str(e)}')
    else:
        form = Carga_alumnos()
        periodos = Periodo.objects.filter(activo=True)
    
    return render(request, 'carga_alumnos.html', {'form': form, 'periodos': periodos})





def crear_periodo(request):
    if request.method == 'POST':
        if 'periodo_id' in request.POST:
            # Manejar la activación del período
            periodo_id = request.POST.get('periodo_id')
            try:
                periodo = Periodo.objects.get(id=periodo_id)
                # Desactivar todos los períodos
                Periodo.objects.update(activo=False)
                # Activar el período seleccionado
                periodo.activo = True
                periodo.save()
                messages.success(request, f'El período {periodo.codigo} ha sido activado.')
            except Periodo.DoesNotExist:
                messages.error(request, 'El período no existe.')
        else:
            # Manejar la creación del período
            codigo = request.POST['codigo']
            fecha_inicio = request.POST['fecha_inicio']
            fecha_fin = request.POST['fecha_fin']
            activo = 'activo' in request.POST  # Convertir el valor del checkbox a booleano

            periodo = Periodo(codigo=codigo, fecha_inicio=fecha_inicio, fecha_fin=fecha_fin, activo=activo)
            try:
                periodo.save()
                messages.success(request, 'Periodo creado exitosamente.')
            except IntegrityError:
                messages.error(request, 'Error: Ya existe un período con el mismo código.')
        return redirect('crear_periodo')

    periodos = Periodo.objects.all()
    return render(request, 'crear_periodo.html', {'periodos': periodos})